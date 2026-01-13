"""Export all location data with RocketReach contacts to Excel."""

import pandas as pd
import time
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import config
from rocketreach_api import RocketReachAPI


def export_to_excel(include_rocketreach=True, only_opportunities=False):
    """
    Export location data to Excel with optional RocketReach contact lookup.

    Args:
        include_rocketreach: Whether to fetch RocketReach contact data
        only_opportunities: If True, only export locations without ATMs
    """
    print("=" * 60)
    print("BITCOIN ATM OPPORTUNITIES - EXCEL EXPORT")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Load location data
    df = pd.read_csv(config.OUTPUT_CSV)
    print(f"\nLoaded {len(df)} locations from database")

    # Filter if only opportunities
    if only_opportunities:
        df = df[df['has_bitcoin_atm'] == False]
        print(f"Filtered to {len(df)} opportunities (no ATM)")

    # Prepare export data
    export_data = []

    # Initialize RocketReach if needed
    rr_api = None
    if include_rocketreach:
        try:
            rr_api = RocketReachAPI()
            status = rr_api.check_api_status()
            credits = next((c for c in status.get('credit_usage', [])
                          if c['credit_type'] == 'premium_lookup'), {})
            print(f"\nRocketReach API connected")
            print(f"Credits remaining: {credits.get('remaining', 'Unknown')}")
        except Exception as e:
            print(f"\nRocketReach API error: {e}")
            print("Continuing without contact lookup...")
            rr_api = None

    print(f"\nProcessing {len(df)} locations...")
    print("-" * 60)

    for idx, row in df.iterrows():
        location_data = {
            'Business Name': row.get('business_name', ''),
            'Business Type': row.get('business_type', ''),
            'Address': row.get('address', ''),
            'Phone': row.get('phone', ''),
            'Google Rating': row.get('google_rating', ''),
            'Has Bitcoin ATM': 'Yes' if row.get('has_bitcoin_atm', False) else 'No',
            'Nearest ATM Distance (km)': row.get('distance_to_nearest_atm', ''),
            'Nearest ATM Operator': row.get('nearest_atm_operator', ''),
            'Opportunity Score': row.get('opportunity_score', ''),
            'Status': row.get('status', 'not_contacted'),
            'Latitude': row.get('latitude', ''),
            'Longitude': row.get('longitude', ''),
            # RocketReach fields
            'Contact 1 - Name': '',
            'Contact 1 - Title': '',
            'Contact 1 - Email': '',
            'Contact 1 - Phone': '',
            'Contact 1 - LinkedIn': '',
            'Contact 2 - Name': '',
            'Contact 2 - Title': '',
            'Contact 2 - Email': '',
            'Contact 2 - Phone': '',
            'Contact 2 - LinkedIn': '',
        }

        # Fetch RocketReach data
        if rr_api and not row.get('has_bitcoin_atm', False):
            try:
                business_name = row.get('business_name', '')
                print(f"[{idx+1}/{len(df)}] Looking up: {business_name[:40]}...", end=" ")

                result = rr_api.get_contact_info(business_name)
                contacts = result.get('contacts', [])

                if contacts:
                    print(f"Found {len(contacts)} contact(s)")
                    # First contact
                    if len(contacts) >= 1:
                        c = contacts[0]
                        location_data['Contact 1 - Name'] = c.get('name', '')
                        location_data['Contact 1 - Title'] = c.get('title', '')
                        location_data['Contact 1 - Email'] = c.get('email', '')
                        location_data['Contact 1 - Phone'] = c.get('phone', '')
                        location_data['Contact 1 - LinkedIn'] = c.get('linkedin', '')
                    # Second contact
                    if len(contacts) >= 2:
                        c = contacts[1]
                        location_data['Contact 2 - Name'] = c.get('name', '')
                        location_data['Contact 2 - Title'] = c.get('title', '')
                        location_data['Contact 2 - Email'] = c.get('email', '')
                        location_data['Contact 2 - Phone'] = c.get('phone', '')
                        location_data['Contact 2 - LinkedIn'] = c.get('linkedin', '')
                else:
                    print("No contacts found")

                # Rate limiting - be gentle with the API
                time.sleep(0.5)

            except Exception as e:
                print(f"Error: {e}")
        else:
            if row.get('has_bitcoin_atm', False):
                print(f"[{idx+1}/{len(df)}] Skipping (has ATM): {row.get('business_name', '')[:40]}")
            elif not rr_api:
                pass  # No API, already handled

        export_data.append(location_data)

    # Create DataFrame
    export_df = pd.DataFrame(export_data)

    # Create Excel file with formatting
    output_file = os.path.join(
        os.path.dirname(config.OUTPUT_CSV),
        f"bitcoin_atm_opportunities_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    print(f"\n{'=' * 60}")
    print("Creating Excel file with formatting...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    # Add data
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Header formatting
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Adjust column widths
    column_widths = {
        'A': 30,  # Business Name
        'B': 18,  # Business Type
        'C': 45,  # Address
        'D': 15,  # Phone
        'E': 12,  # Rating
        'F': 15,  # Has ATM
        'G': 20,  # Distance
        'H': 20,  # Operator
        'I': 12,  # Score
        'J': 15,  # Status
        'K': 12,  # Lat
        'L': 12,  # Lng
        'M': 25,  # Contact 1 Name
        'N': 25,  # Contact 1 Title
        'O': 30,  # Contact 1 Email
        'P': 18,  # Contact 1 Phone
        'Q': 35,  # Contact 1 LinkedIn
        'R': 25,  # Contact 2 Name
        'S': 25,  # Contact 2 Title
        'T': 30,  # Contact 2 Email
        'U': 18,  # Contact 2 Phone
        'V': 35,  # Contact 2 LinkedIn
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Add conditional formatting for Has ATM column
    for row in range(2, len(export_data) + 2):
        cell = ws.cell(row=row, column=6)  # Has ATM column
        if cell.value == "Yes":
            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save
    wb.save(output_file)

    print(f"\n{'=' * 60}")
    print("EXPORT COMPLETE!")
    print("=" * 60)
    print(f"\nFile saved: {output_file}")
    print(f"Total records: {len(export_data)}")

    # Count contacts found
    contacts_found = sum(1 for d in export_data if d.get('Contact 1 - Email') or d.get('Contact 1 - Name'))
    print(f"Locations with contacts: {contacts_found}")

    return output_file


if __name__ == "__main__":
    import sys

    # Check for command line args
    only_opps = "--opportunities-only" in sys.argv or "-o" in sys.argv
    no_rr = "--no-rocketreach" in sys.argv or "-n" in sys.argv

    print("\nExport Options:")
    print(f"  Include RocketReach: {'No' if no_rr else 'Yes'}")
    print(f"  Only Opportunities: {'Yes' if only_opps else 'No (all locations)'}")
    print()

    export_to_excel(
        include_rocketreach=not no_rr,
        only_opportunities=only_opps
    )
